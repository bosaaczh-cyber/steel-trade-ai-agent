# -*- coding: utf-8 -*-
"""
Abter Steel — DirectMail 发送统计模块
══════════════════════════════════════
对接阿里云邮件推送API，追踪每封邮件的发送/打开/点击/退信数据

运行方式：
  python steel_stats.py --mode today    # 今日发送统计
  python steel_stats.py --mode week     # 本周统计
  python steel_stats.py --mode export   # 导出Excel报表
  python steel_stats.py --mode bounce   # 查看退信地址（无效邮箱）
"""

import os
import json
import csv
import hmac
import base64
import hashlib
import uuid
import argparse
from datetime import datetime, date, timedelta
from urllib import parse, request as urlrequest
from pathlib import Path

from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

load_dotenv()

# ══════════════════════════════════════════
# 配置
# ══════════════════════════════════════════
ACCESS_KEY_ID     = os.environ.get("ALIYUN_ACCESS_KEY_ID", "")
ACCESS_KEY_SECRET = os.environ.get("ALIYUN_ACCESS_KEY_SECRET", "")
FROM_ADDRESS      = os.environ.get("ALIYUN_FROM_ADDRESS", "")
API_ENDPOINT      = "https://dm.aliyuncs.com"

# ══════════════════════════════════════════
# 阿里云 API 签名（复用 Steel sender.py 逻辑）
# ══════════════════════════════════════════
def _percent_encode(s: str) -> str:
    return parse.quote(str(s), safe="")

def _sign(params: dict, secret: str) -> str:
    sorted_params = sorted(params.items())
    query_string  = "&".join([f"{_percent_encode(k)}={_percent_encode(v)}"
                               for k, v in sorted_params])
    string_to_sign = f"GET&{_percent_encode('/')}&{_percent_encode(query_string)}"
    key    = (secret + "&").encode("utf-8")
    hashed = hmac.new(key, string_to_sign.encode("utf-8"), hashlib.sha1)
    return base64.b64encode(hashed.digest()).decode("utf-8")

def _call_api(action: str, extra_params: dict) -> dict:
    params = {
        "Format":           "JSON",
        "Version":          "2015-11-23",
        "AccessKeyId":      ACCESS_KEY_ID,
        "SignatureMethod":  "HMAC-SHA1",
        "Timestamp":        datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ"),
        "SignatureVersion": "1.0",
        "SignatureNonce":   str(uuid.uuid4()),
        "Action":           action,
        **extra_params,
    }
    params["Signature"] = _sign(params, ACCESS_KEY_SECRET)
    url = f"{API_ENDPOINT}/?{parse.urlencode(params)}"
    try:
        with urlrequest.urlopen(url, timeout=15) as resp:
            return json.loads(resp.read().decode("utf-8"))
    except Exception as e:
        return {"Error": str(e)}

# ══════════════════════════════════════════
# 阿里云 DirectMail 统计 API
# ══════════════════════════════════════════
def get_send_statistics(start_date: str, end_date: str) -> dict:
    """
    查询指定日期范围的发送统计
    start_date/end_date 格式：YYYY-MM-DD
    返回：{发送量, 送达量, 打开量, 点击量, 退信量, 垃圾举报量}
    """
    result = _call_api("GetMailAddressMsgStatByParam", {
        "AccountName": FROM_ADDRESS,
        "StartTime":   start_date,
        "EndTime":     end_date,
    })
    return result


def get_bounce_list() -> list:
    """
    获取退信地址列表（无效邮箱）
    这些地址应该从发信列表中剔除
    """
    result = _call_api("QueryInvalidAddress", {
        "AccountName": FROM_ADDRESS,
    })
    addresses = result.get("InvalidAddressList", {}).get("Invalid", [])
    return addresses


def get_send_detail(start_date: str, end_date: str, page: int = 1) -> dict:
    """
    查询逐封邮件的发送明细
    """
    result = _call_api("SenderStatisticsMailQueueByParam", {
        "AccountName": FROM_ADDRESS,
        "StartTime":   start_date,
        "EndTime":     end_date,
        "Offset":      str((page - 1) * 10),
        "Length":      "10",
    })
    return result


def get_quota_status() -> dict:
    """查询今日剩余发信额度"""
    result = _call_api("DescQuota", {})
    return result


# ══════════════════════════════════════════
# 本地日志统计（不依赖API，用send_log.json）
# ══════════════════════════════════════════
def local_stats_from_log(days: int = 7) -> dict:
    """
    从本地 send_log.json 统计发送数据
    不需要API，离线也能用
    """
    log_file = Path("send_log.json")
    if not log_file.exists():
        return {}

    with open(log_file, "r", encoding="utf-8") as f:
        log = json.load(f)

    cutoff = date.today() - timedelta(days=days)
    recent = [l for l in log if date.fromisoformat(l.get("date", "2000-01-01")) >= cutoff]

    # 按日期分组
    by_date = {}
    for entry in recent:
        d = entry.get("date", "unknown")
        if d not in by_date:
            by_date[d] = {"total": 0, "success": 0, "failed": 0, "companies": []}
        by_date[d]["total"]   += 1
        by_date[d][entry.get("status", "failed")] += 1
        if entry.get("company"):
            by_date[d]["companies"].append(entry["company"])

    # 汇总
    total   = len(recent)
    success = sum(1 for l in recent if l.get("status") == "success")
    failed  = sum(1 for l in recent if l.get("status") == "failed")

    return {
        "period_days":   days,
        "total":         total,
        "success":       success,
        "failed":        failed,
        "success_rate":  f"{success/total*100:.1f}%" if total else "0%",
        "by_date":       by_date,
        "all_logs":      recent,
    }


# ══════════════════════════════════════════
# 打印统计表
# ══════════════════════════════════════════
def print_stats(stats: dict):
    if not stats:
        print("  ⚠️  暂无发送记录")
        return

    print(f"\n  {'─'*55}")
    print(f"  📊 最近 {stats['period_days']} 天发送统计")
    print(f"  {'─'*55}")
    print(f"  总发送量：{stats['total']} 封")
    print(f"  成功送达：{stats['success']} 封")
    print(f"  发送失败：{stats['failed']} 封")
    print(f"  成功率：  {stats['success_rate']}")
    print(f"\n  按日期明细：")
    print(f"  {'日期':<12} {'发送':<6} {'成功':<6} {'失败':<6} {'公司列表（前3）'}")
    print(f"  {'─'*55}")
    for d in sorted(stats["by_date"].keys(), reverse=True):
        day_data = stats["by_date"][d]
        companies = ", ".join(day_data["companies"][:3])
        if len(day_data["companies"]) > 3:
            companies += f" 等{len(day_data['companies'])}家"
        print(f"  {d:<12} {day_data['total']:<6} {day_data['success']:<6} {day_data['failed']:<6} {companies}")
    print(f"  {'─'*55}\n")


# ══════════════════════════════════════════
# 导出 Excel 报表
# ══════════════════════════════════════════
def export_report(stats: dict, filename: str = ""):
    if not filename:
        filename = f"email_stats_{date.today()}.xlsx"

    wb = Workbook()

    # ── Sheet1：汇总 ──────────────────────
    ws1 = wb.active
    ws1.title = "发送汇总"

    # 标题
    ws1.merge_cells("A1:F1")
    ws1["A1"] = f"Abter Steel — 邮件发送统计报表（{date.today()}）"
    ws1["A1"].font      = Font(bold=True, size=13, color="FFFFFF")
    ws1["A1"].fill      = PatternFill("solid", start_color="1A3C5E", end_color="1A3C5E")
    ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 28

    # 汇总数字
    summary_data = [
        ("统计周期", f"最近 {stats.get('period_days', 7)} 天"),
        ("总发送量", stats.get("total", 0)),
        ("成功送达", stats.get("success", 0)),
        ("发送失败", stats.get("failed", 0)),
        ("成功率",   stats.get("success_rate", "0%")),
    ]
    for i, (label, value) in enumerate(summary_data, 2):
        ws1[f"A{i}"] = label
        ws1[f"B{i}"] = value
        ws1[f"A{i}"].font = Font(bold=True)

    ws1.column_dimensions["A"].width = 15
    ws1.column_dimensions["B"].width = 20

    # ── Sheet2：日期明细 ──────────────────
    ws2 = wb.create_sheet("每日明细")
    headers = ["日期", "发送量", "成功", "失败", "成功率", "发送公司"]
    for col, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color="2E86AB", end_color="2E86AB")
        cell.alignment = Alignment(horizontal="center")

    ws2.column_dimensions["A"].width = 14
    ws2.column_dimensions["F"].width = 50

    for row_idx, (d, data) in enumerate(
        sorted(stats.get("by_date", {}).items(), reverse=True), 2
    ):
        total_day   = data["total"]
        success_day = data["success"]
        rate        = f"{success_day/total_day*100:.1f}%" if total_day else "0%"
        companies   = ", ".join(data["companies"])

        row_data = [d, total_day, success_day, data["failed"], rate, companies]
        for col, val in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col, value=val)
            cell.alignment = Alignment(vertical="center")

    # ── Sheet3：完整发送记录 ──────────────
    ws3 = wb.create_sheet("完整记录")
    log_headers = ["日期", "时间", "公司", "联系人", "邮箱", "主题", "状态", "RequestID"]
    for col, h in enumerate(log_headers, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color="1A3C5E", end_color="1A3C5E")
        ws3.column_dimensions[get_column_letter(col)].width = 18

    ws3.column_dimensions["E"].width = 35
    ws3.column_dimensions["F"].width = 50

    for row_idx, entry in enumerate(stats.get("all_logs", []), 2):
        status_color = "D6F0D6" if entry.get("status") == "success" else "FFE0E0"
        row_data = [
            entry.get("date", ""),
            entry.get("time", ""),
            entry.get("company", ""),
            entry.get("contact", ""),
            entry.get("email", ""),
            entry.get("subject", "")[:40],
            "✅ 成功" if entry.get("status") == "success" else "❌ 失败",
            entry.get("request_id", ""),
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws3.cell(row=row_idx, column=col, value=val)
            cell.fill = PatternFill("solid", start_color=status_color,
                                     end_color=status_color)
            cell.alignment = Alignment(vertical="center")

    wb.save(filename)
    print(f"  📄 报表已导出 → {filename}")
    return filename


# ══════════════════════════════════════════
# 退信处理：从发送列表中自动剔除无效邮箱
# ══════════════════════════════════════════
def clean_bounce_from_csv(csv_path: str) -> str:
    """
    读取退信列表，从CSV发信名单中剔除无效邮箱
    返回清洗后的CSV文件路径
    """
    print("\n  🔍 查询退信列表...")

    # 先从本地日志找失败记录
    log_file = Path("send_log.json")
    failed_emails = set()
    if log_file.exists():
        with open(log_file, "r", encoding="utf-8") as f:
            log = json.load(f)
        failed_emails = {l["email"] for l in log
                        if l.get("status") == "failed" and l.get("email")}

    # 再从阿里云API查退信
    bounces = get_bounce_list()
    for b in bounces:
        addr = b.get("Address", "")
        if addr:
            failed_emails.add(addr.lower())

    if not failed_emails:
        print("  ✅ 没有退信记录，名单干净")
        return csv_path

    print(f"  ⚠️  发现 {len(failed_emails)} 个无效邮箱，正在清洗...")

    # 读取原CSV
    rows = []
    with open(csv_path, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames or []
        rows = list(reader)

    # 过滤
    clean_rows  = [r for r in rows
                   if r.get("收件人邮箱", "").lower() not in failed_emails]
    removed     = len(rows) - len(clean_rows)

    # 保存清洗后的文件
    clean_path = csv_path.replace(".csv", "_cleaned.csv")
    with open(clean_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(clean_rows)

    print(f"  ✅ 清洗完成：剔除 {removed} 个无效邮箱 → {clean_path}")
    return clean_path


# ══════════════════════════════════════════
# 主入口
# ══════════════════════════════════════════
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Abter Steel 发送统计")
    parser.add_argument(
        "--mode",
        choices=["today", "week", "month", "export", "bounce", "quota"],
        default="week",
        help="today=今日 | week=本周 | month=本月 | export=导出Excel | bounce=退信 | quota=查额度",
    )
    args = parser.parse_args()

    print("=" * 55)
    print("  Abter Steel — DirectMail 发送统计")
    print(f"  模式：{args.mode.upper()}  |  今天：{date.today()}")
    print("=" * 55)

    if args.mode == "today":
        stats = local_stats_from_log(days=1)
        print_stats(stats)

    elif args.mode == "week":
        stats = local_stats_from_log(days=7)
        print_stats(stats)

    elif args.mode == "month":
        stats = local_stats_from_log(days=30)
        print_stats(stats)

    elif args.mode == "export":
        stats = local_stats_from_log(days=30)
        print_stats(stats)
        if stats:
            export_report(stats)

    elif args.mode == "bounce":
        print("\n  🔍 查询退信/无效邮箱列表...")
        if not ACCESS_KEY_SECRET:
            print("  ⚠️  未配置 ALIYUN_ACCESS_KEY_SECRET，使用本地失败记录")
            log_file = Path("send_log.json")
            if log_file.exists():
                with open(log_file) as f:
                    log = json.load(f)
                failed = [l for l in log if l.get("status") == "failed"]
                if failed:
                    print(f"\n  本地失败记录（{len(failed)} 条）：")
                    for l in failed:
                        print(f"  · {l.get('company',''):<25} {l.get('email','')}")
                else:
                    print("  ✅ 没有失败记录")
        else:
            bounces = get_bounce_list()
            if bounces:
                print(f"\n  退信地址（{len(bounces)} 个）：")
                for b in bounces:
                    print(f"  · {b.get('Address', '')}  原因：{b.get('Reason', '')}")
            else:
                print("  ✅ 没有退信记录")

    elif args.mode == "quota":
        print("\n  🔍 查询今日发信额度...")
        if not ACCESS_KEY_SECRET:
            print("  ⚠️  未配置 ALIYUN_ACCESS_KEY_SECRET")
        else:
            result = get_quota_status()
            if "Error" not in result:
                print(f"  日额度：{result.get('DayQuota', '未知')}")
                print(f"  已用：  {result.get('DayQuotaUsed', '未知')}")
                remaining = int(result.get('DayQuota', 0)) - int(result.get('DayQuotaUsed', 0))
                print(f"  剩余：  {remaining}")
            else:
                print(f"  查询失败：{result}")